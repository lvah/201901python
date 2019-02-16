"""
文件名: $NAME.py
日期: 16  
作者: lvah
联系: xc_guofan@qq.com
代码描述: 



url = "http://movie.douban.com/top250/"

需要获取的信息: 电影名称， 电影评分， 评价人数， 电影短评
"""
import re

import requests
from bs4 import BeautifulSoup

def get_content(url,):
    try:
        user_agent = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.109 Safari/537.36"
        response = requests.get(url,  headers={'User-Agent': user_agent})
        response.raise_for_status()   # 如果返回的状态码不是200， 则抛出异常;
        response.encoding = response.apparent_encoding  # 判断网页的编码格式， 便于respons.text知道如何解码;
    except Exception as e:
        print("爬取错误")
    else:

        print(response.url)
        print("爬取成功!")
        return  response.content

def parser_content(htmlContent):
    # 实例化soup对象， 便于处理；
    soup = BeautifulSoup(htmlContent, 'html.parser')
    #  1). 电影信息存储在ol标签里面的li标签:
    #  <ol class="grid_view">
    olObj = soup.find_all('ol', class_='grid_view')[0]

    #  2). 获取每个电影的详细信息, 存储在li标签;
    details = olObj.find_all('li')


    for detail in details:
        # #  3). 获取电影名称;
        movieName = detail.find('span', class_='title').get_text()

        # 4). 电影评分：
        movieScore = detail.find('span', class_='rating_num').get_text()

        # 5). 评价人数***************
        # 必须要转换类型为字符串
        movieCommentNum = str(detail.find(text=re.compile('\d+人评价')))


        # 6). 电影短评
        movieCommentObj = detail.find('span', class_='inq')
        if movieCommentObj:
            movieComment = movieCommentObj.get_text()
        else:
            movieComment = "无短评"

        movieInfo.append((movieName, movieScore, movieCommentNum, movieComment))




import openpyxl


def create_to_excel(wbname, data, sheetname='Sheet1', ):
    """
    将制定的信息保存到新建的excel表格中;

    :param wbname:
    :param data: 往excel中存储的数据;
    :param sheetname:
    :return:
    """

    print("正在创建excel表格%s......" % (wbname))

    # wb = openpyxl.load_workbook(wbname)
    #  如果文件不存在， 自己实例化一个WorkBook的对象;
    wb = openpyxl.Workbook()
    # 获取当前活动工作表的对象
    sheet = wb.active
    # 修改工作表的名称
    sheet.title = sheetname
    # 将数据data写入excel表格中;
    print("正在写入数据........")
    for row, item in enumerate(data):  # data发现有4行数据， item里面有三列数据;
        print(item)
        for column, cellValue in enumerate(item):

            # cell = sheet.cell(row=row + 1, column=column + 1, value=cellValue)
            cell = sheet.cell(row=row+1, column=column + 1)
            cell.value = cellValue

    wb.save(wbname)
    print("保存工作薄%s成功......." % (wbname))


if __name__ == '__main__':
    doubanTopPage = 2
    perPage = 25
    # [(), (), ()]
    movieInfo = []
    # 1, 2, 3 ,4, 5
    for page in range(1, doubanTopPage+1):
        # start的值= (当前页-1)*每页显示的数量(25)
        url = "https://movie.douban.com/top250?start=%s" %((page-1)*perPage)
        content = get_content(url)
        parser_content(content)


    create_to_excel('/tmp/hello.xlsx', movieInfo, sheetname="豆瓣电影信息")


