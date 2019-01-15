"""
文件名: .py
创建时间: 2019-01-15 16:
作者: lvah
联系方式: 976131979@qq.com
代码描述:


# 1. Excel的基本定义
- 工作薄(workbook)：一个Excel电子表格；
- 工作表(sheet): 每个工作薄里面包含的表， 默认有3个;
- 活动表(active sheet): 用户当前查看的工作表;
- 单元格(cell): 特定的行和列构成的格子;
- 列(column): 列地址是从A开始的；
- 行(row):行地址是从1开始的；




# 2. 版本 < 2007使用xlrd模块实现
# import xlrd
# wb = xlrd.open_workbook('/tmp/Book1.xls')
# print(wb.sheet_names())


"""

import openpyxl

# 1. 导入工作薄
wb = openpyxl.load_workbook('/tmp/Book1.xlsx')
# 2. 查看包含的工作表
print(wb.sheetnames)
# 3. 查看当前活动表
print(wb.active)
# 4. 从工作薄中选择要操作的工作表
sheet = wb['修改的工作表']
# 5. 查看制定单元格的数据
# cell = sheet['B3']
cell = sheet.cell(row=3, column=2)
# 6. 获取单元格详细信息
print(cell.value)

# 7. 修改单元格信息
cell.value = 1111111

# 8. 获取工作表的总行和总列数
print(sheet.max_column)
print(sheet.max_row)


# 9. 修改单元表的名称
print(sheet.title)
sheet.title = "修改的工作表"


# 10. 遍历历EXCEL表格的所有数据；
print(sheet.rows)   # 返回的是一个生成器
# 遍历历每一行
for row in sheet.rows:
    # 获取每一行， 每个单元格的数据
    for cell in row:
        print(cell.value, end='\t')
    print('\n')


# 11. 保存修改的信息
wb.save('/tmp/Book1.xlsx')
