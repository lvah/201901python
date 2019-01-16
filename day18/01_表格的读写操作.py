"""
文件名: .py
创建时间: 2019-01-16 09:
作者: lvah
联系方式: 976131979@qq.com
代码描述:


"""

import openpyxl


def create_to_excel(wbname, data, sheetname='Sheet1', ):
    """
    将制定的信息保存到新建的excel表格中;

    :param wbname:
    :param data: 往excel中存储的数据;
    :param sheetname:
    :return:
    """

    print("正在创建excel表格%s......" % (wbname))

    # wb = openpyxl.load_workbook(wbname)

    #  如果文件不存在， 自己实例化一个WorkBook的对象;
    wb = openpyxl.Workbook()

    # 获取当前活动工作表的对象
    sheet = wb.active

    # 修改工作表的名称
    sheet.title = sheetname

    # 将数据data写入excel表格中;
    print("正在写入数据........")
    for row, item in enumerate(data):  # data发现有4行数据， item里面有三列数据;
        for column, cellValue in enumerate(item):
            cell = sheet.cell(row=row + 1, column=column + 1, value=cellValue)
            # cell = sheet.cell(row=row+1, column=column + 1)
            # cell.value = cellValue

    wb.save(wbname)
    print("保存工作薄%s成功......." % (wbname))


def readwb(wbname, sheetname=None):
    # 1. 加载工作薄
    wb = openpyxl.load_workbook(filename=wbname)

    # 2. 选择操作的工作表
    if not sheetname:
        sheet = wb.active
    else:
        sheet = wb[sheetname]

    #  读取数据， 存储为python的数据结构
    goodsInfo = []
    for row in sheet.rows:
        rowValues = [cell.value for cell in row]
        goodsInfo.append(rowValues)
    return goodsInfo


if __name__ == '__main__':
    data = [
        ['书记名称', '数量', '价格'],
        ['python核心编程', 60, 90],
        ['Java核心编程', 50, 120],
        ['Php核心编程', 100, 80],
    ]

    print("1. " + "*" * 50)
    create_to_excel('doc/excel01.xlsx', data, '书籍信息统计')

    goodsInfo = readwb('doc/excel01.xlsx', '书籍信息统计')
    # print(goodsInfo)

    print("2. " + "*" * 50)
    # 按照商品数量进行排序 sorted()
    # ['书记名称', '数量', '价格'] + [['python核心编程', 60, 90],
    #                         ['Java核心编程', 50, 120],
    #                         ['Php核心编程', 100, 80], ]

    # [
    # ['书记名称', '数量', '价格'],
    #  ['python核心编程', 60, 90],
    #  ['Java核心编程', 50, 120],
    #  ['Php核心编程', 100, 80],
    #
    #
    # ]
    numSortInfo = [goodsInfo[0]]+  sorted(goodsInfo[1:], key=lambda x: x[1])
    print(numSortInfo)
    create_to_excel('doc/sorted_by_num.xlsx', numSortInfo, '书籍信息统计按照数量排序')

    print("3. " + "*" * 50)
    #  按照商品价格进行排序 sorted()
    priceSortInfo = [goodsInfo[0]] + sorted(goodsInfo[1:], key=lambda x: x[1])
    print(priceSortInfo)
    create_to_excel('doc/sorted_by_price.xlsx', numSortInfo, '书籍信息统计按照价格排序')
